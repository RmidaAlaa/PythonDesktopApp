"""Excel report generation."""

import platform
from pathlib import Path
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QCoreApplication
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except Exception:
    import re
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

from .config import Config
from .device_detector import Device
from .logger import setup_logger

logger = setup_logger("ReportGenerator")


class ReportGenerator:
    """Generates Excel reports."""
    
    def __init__(self):
        self.logger = logger
        self._invalid_sheet_chars = set('[]:*?/\\')
    
    def generate_report(self, devices: List[Device], operator_info: Dict, 
                       machine_type: str, machine_id: str) -> Path:
        """Generate an Excel report."""
        workbook = Workbook()
        
        # Create metadata sheet
        self._create_metadata_sheet(workbook, operator_info)
        
        # Create devices sheet
        self._create_devices_sheet(workbook, devices, machine_type, machine_id)
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"AWG_Kumulus_Report_{timestamp}.xlsx"
        report_path = Config.APPDATA_DIR / filename
        
        workbook.save(report_path)
        logger.info(f"Report saved to {report_path}")
        
        return report_path
    
    def _create_metadata_sheet(self, workbook: Workbook, operator_info: Dict):
        """Create the metadata sheet."""
        sheet = workbook.active
        sheet.title = self._safe_sheet_title(QCoreApplication.translate("ReportGenerator", "Metadata"))
        
        # Get PC info
        pc_name = platform.node()
        pc_os = f"{platform.system()} {platform.release()}"
        
        headers = [
            QCoreApplication.translate("ReportGenerator", "Property"),
            QCoreApplication.translate("ReportGenerator", "Value")
        ]
        data = [
            [QCoreApplication.translate("ReportGenerator", "Timestamp"), datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            [QCoreApplication.translate("ReportGenerator", "Application"), "AWG Kumulus Device Manager"],
            [QCoreApplication.translate("ReportGenerator", "Version"), "1.0.0"],
            [QCoreApplication.translate("ReportGenerator", "Operator Name"), self._sanitize_cell_value(operator_info.get("name", "N/A"))],
            [QCoreApplication.translate("ReportGenerator", "Operator Email"), self._sanitize_cell_value(operator_info.get("email", "N/A"))],
            [QCoreApplication.translate("ReportGenerator", "Client Name"), self._sanitize_cell_value(operator_info.get("client_name", "N/A"))],
            [QCoreApplication.translate("ReportGenerator", "Machine Type"), self._sanitize_cell_value(operator_info.get("machine_type", "N/A"))],
            [QCoreApplication.translate("ReportGenerator", "Machine ID"), self._sanitize_cell_value(operator_info.get("machine_id", "N/A"))],
            [QCoreApplication.translate("ReportGenerator", "PC Name"), self._sanitize_cell_value(pc_name)],
            [QCoreApplication.translate("ReportGenerator", "PC OS"), self._sanitize_cell_value(pc_os)],
            [QCoreApplication.translate("ReportGenerator", "Platform"), self._sanitize_cell_value(platform.machine())],
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(1, col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write data
        for row, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                sheet.cell(row, col, self._sanitize_cell_value(value))
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 40
    
    def _create_devices_sheet(self, workbook: Workbook, devices: List[Device], 
                             machine_type: str, machine_id: str):
        """Create the devices sheet."""
        sheet = workbook.create_sheet(title=self._safe_sheet_title(QCoreApplication.translate("ReportGenerator", "Devices")))
        
        headers = [
            QCoreApplication.translate("ReportGenerator", "Machine Type"), 
            QCoreApplication.translate("ReportGenerator", "Machine ID"), 
            QCoreApplication.translate("ReportGenerator", "Board Type"), 
            QCoreApplication.translate("ReportGenerator", "Port"), 
            QCoreApplication.translate("ReportGenerator", "VID"), 
            QCoreApplication.translate("ReportGenerator", "PID"), 
            QCoreApplication.translate("ReportGenerator", "UID"), 
            QCoreApplication.translate("ReportGenerator", "Chip ID"), 
            QCoreApplication.translate("ReportGenerator", "MAC Address"), 
            QCoreApplication.translate("ReportGenerator", "Manufacturer"), 
            QCoreApplication.translate("ReportGenerator", "Serial Number"), 
            QCoreApplication.translate("ReportGenerator", "Firmware Version"), 
            QCoreApplication.translate("ReportGenerator", "Hardware Version"), 
            QCoreApplication.translate("ReportGenerator", "Flash Size"), 
            QCoreApplication.translate("ReportGenerator", "CPU Frequency"), 
            QCoreApplication.translate("ReportGenerator", "Timestamp")
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(1, col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Write device data
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for row, device in enumerate(devices, 2):
            data = [
                self._sanitize_cell_value(machine_type),
                self._sanitize_cell_value(machine_id),
                self._sanitize_cell_value(device.board_type.value),
                self._sanitize_cell_value(device.port),
                self._sanitize_cell_value(f"0x{device.vid:04X}") if device.vid else self._sanitize_cell_value("N/A"),
                self._sanitize_cell_value(f"0x{device.pid:04X}") if device.pid else self._sanitize_cell_value("N/A"),
                self._sanitize_cell_value(device.uid or "N/A"),
                self._sanitize_cell_value(device.chip_id or "N/A"),
                self._sanitize_cell_value(device.mac_address or "N/A"),
                self._sanitize_cell_value(device.manufacturer or "N/A"),
                self._sanitize_cell_value(device.serial_number or "N/A"),
                self._sanitize_cell_value(device.firmware_version or "N/A"),
                self._sanitize_cell_value(device.hardware_version or "N/A"),
                self._sanitize_cell_value(device.flash_size or "N/A"),
                self._sanitize_cell_value(device.cpu_frequency or "N/A"),
                timestamp
            ]
            
            for col, value in enumerate(data, 1):
                sheet.cell(row, col, value)
        
        # Freeze top row
        sheet.freeze_panes = 'A2'
        
        # Adjust column widths
        column_widths = [15, 15, 12, 10, 10, 10, 20, 15, 18, 20, 20, 15, 15, 12, 12, 20]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width

    def _sanitize_cell_value(self, value):
        """Sanitize value for Excel cell."""
        if value is None:
            return QCoreApplication.translate("ReportGenerator", "N/A")
        if isinstance(value, str):
            if not value or value == "N/A":
                return QCoreApplication.translate("ReportGenerator", "N/A")
            # Remove illegal characters
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value

    def _safe_sheet_title(self, title: str, fallback: str = "Sheet") -> str:
        """Ensure sheet title is valid for Excel."""
        if not title:
            return QCoreApplication.translate("ReportGenerator", fallback)
        
        # Remove invalid characters: \ / ? * [ ] :
        title = "".join(c for c in title if c not in r"\/?*[]:")
        
        # Max length 31
        title = title[:31]
        title = title.strip()
        
        if not title:
            return QCoreApplication.translate("ReportGenerator", fallback)
            
        return title

